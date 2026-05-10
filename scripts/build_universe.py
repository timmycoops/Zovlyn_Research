#!/usr/bin/env python3
"""
Build drill_universe.json from the uploaded Excel of drill-hole results.

Inputs:
  data/Drill_Holes.xlsx  (copied from upload)

Outputs:
  drill_universe.json  with:
    - meta: source / counts / generated_at
    - intercepts: compact [project, commodity_group, width_m, auEq_gpt] for all valid rows
    - kde_contours: per-commodity Gaussian KDE polygons (50% and 80% mass enclosure) on
      log10(true_width) x log10(AuEq grade) — true width assumed = downhole width.
    - projects: project_name -> {n_intercepts, primary_commodity}

KDE method:
  * scipy.stats.gaussian_kde on stacked (log10 w, log10 g) with default bw.
  * Evaluated on an 80x80 grid spanning log10([1,500]) x log10([0.3,3000]).
  * Contour level for X% mass enclosure = density threshold k such that the
    integral of f >= k is X% of total mass; computed via sorted cumulative sum.
"""

import json
import math
import datetime as dt
from collections import Counter, defaultdict

import openpyxl
import numpy as np
from scipy.stats import gaussian_kde
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

XLSX = '/home/user/Zovlyn_Research/data/Drill_Holes.xlsx'
OUT  = '/home/user/Zovlyn_Research/drill_universe.json'

# Same axis bounds as the chart (drill_intercept.html: X_MIN..X_MAX, Y_MIN..Y_MAX)
LOG_X = (math.log10(1), math.log10(500))
LOG_Y = (math.log10(0.3), math.log10(3000))

# Commodity grouping — collapse the long tail
PRIME_TO_GROUP = {
    'Au': 'Au', 'Ag': 'Au', 'Pd': 'Au', 'Pt': 'Au', 'Platinium': 'Au',
    'Cu': 'Cu', 'Mo': 'Cu',
    'Li2O': 'Li', 'Sn': 'Li',
    'Ni': 'Ni', 'Co': 'Ni',
    'U': 'U',
    'Zn': 'Zn', 'Pb': 'Zn',
}


def load_records():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    records = []
    for r in rows[1:]:
        rec = dict(zip(header, r))
        w = rec.get('Width')
        g = rec.get('AuEq (G/t)')
        if not isinstance(w, (int, float)) or not isinstance(g, (int, float)):
            continue
        if w <= 0 or g <= 0:
            continue
        proj = (rec.get('Project Name') or '').strip()
        prime = (rec.get('Prime Metal') or '').strip()
        group = PRIME_TO_GROUP.get(prime, 'Other')
        records.append({
            'p': proj or None,
            'c': group,
            'pm': prime or None,
            'w': round(float(w), 2),
            'g': round(float(g), 3),
            'co': (rec.get('Company Code') or '').strip() or None,
            'l': (rec.get('Location') or '').strip() or None,
        })
    return records


def mass_level(Z, fraction):
    flat = Z.flatten()
    if not flat.any():
        return None
    sort_desc = np.sort(flat)[::-1]
    cum = np.cumsum(sort_desc)
    cum /= cum[-1]
    idx = int(np.searchsorted(cum, fraction))
    idx = min(idx, len(sort_desc) - 1)
    return float(sort_desc[idx])


def simplify_polygon(poly, max_pts=60):
    """Uniform downsample so polygons stay light in JSON."""
    if len(poly) <= max_pts:
        return poly
    step = len(poly) / max_pts
    out = [poly[int(i * step)] for i in range(max_pts)]
    out.append(poly[-1])
    return out


def kde_contours_for(records, group_name):
    pts = [(math.log10(r['w']), math.log10(r['g'])) for r in records if r['c'] == group_name]
    n = len(pts)
    if n < 25:
        return None
    xs = np.array([p[0] for p in pts])
    ys = np.array([p[1] for p in pts])
    kde = gaussian_kde(np.vstack([xs, ys]))
    xg = np.linspace(LOG_X[0], LOG_X[1], 80)
    yg = np.linspace(LOG_Y[0], LOG_Y[1], 80)
    X, Y = np.meshgrid(xg, yg)
    grid = np.vstack([X.ravel(), Y.ravel()])
    Z = kde(grid).reshape(X.shape)
    levels = {
        '50': mass_level(Z, 0.50),
        '80': mass_level(Z, 0.80),
    }
    levels = {k: v for k, v in levels.items() if v is not None}
    if not levels:
        return None
    out = {'n': n, 'contours': {}}
    fig, ax = plt.subplots()
    sorted_levels = sorted(set(levels.values()))
    cs = ax.contour(X, Y, Z, levels=sorted_levels)
    plt.close(fig)
    # cs.allsegs is list of segments per level (in matplotlib >=3.8 use cs.get_paths())
    if hasattr(cs, 'allsegs'):
        for level_idx, level_val in enumerate(cs.levels):
            label = next((k for k, v in levels.items() if abs(v - level_val) < 1e-12), None)
            if label is None:
                continue
            polys = []
            for seg in cs.allsegs[level_idx]:
                if len(seg) < 4:
                    continue
                poly = [[round(10**x, 3), round(10**y, 4)] for x, y in seg]
                polys.append(simplify_polygon(poly))
            if polys:
                out['contours'][label] = polys
    return out if out['contours'] else None


def build():
    records = load_records()
    print(f'Loaded {len(records)} valid intercepts')

    # Compact intercepts
    intercepts_arr = []
    for r in records:
        intercepts_arr.append([r['p'] or '', r['c'], r['w'], r['g']])

    # Project rollup
    proj_groups = defaultdict(list)
    for r in records:
        if r['p']:
            proj_groups[r['p']].append(r)
    projects = []
    for name, recs in proj_groups.items():
        primary = Counter(r['c'] for r in recs).most_common(1)[0][0]
        projects.append({'name': name, 'n': len(recs), 'primary': primary})
    projects.sort(key=lambda p: -p['n'])

    # Per-commodity KDE
    print('Computing per-commodity KDE…')
    commodities_present = sorted(set(r['c'] for r in records))
    kde_out = {}
    for c in commodities_present:
        result = kde_contours_for(records, c)
        if result:
            kde_out[c] = result
            print(f'  {c}: n={result["n"]}, levels={list(result["contours"].keys())}, '
                  f'polys total={sum(len(v) for v in result["contours"].values())}')

    payload = {
        'meta': {
            'source': 'User upload — 1ef9eef6-Drill_Holes.xlsx',
            'generated_at': dt.datetime.utcnow().isoformat() + 'Z',
            'n_intercepts': len(records),
            'n_projects': len(projects),
            'true_width_assumption': '100% (downhole length used as true width)',
            'commodity_grouping': PRIME_TO_GROUP,
            'kde': {
                'method': 'scipy.stats.gaussian_kde, default bandwidth',
                'space': 'log10(true_width_m) x log10(AuEq_gpt)',
                'grid': '80x80 over log10([1,500]) x log10([0.3,3000])',
                'levels': '50% / 80% probability-mass enclosure',
            },
        },
        'projects': projects,
        'intercepts': intercepts_arr,
        'kde_contours': kde_out,
    }

    with open(OUT, 'w') as f:
        json.dump(payload, f, separators=(',', ':'))
    import os
    sz = os.path.getsize(OUT)
    print(f'\nWrote {OUT}  ({sz/1024:.1f} kB)')


if __name__ == '__main__':
    build()
